# -*- coding: utf-8 -*-
"""
bulk_validation.py
──────────────────────────────────────────────────────────────────────────
مرحلهٔ ۳ از نقشهٔ راه: پیش‌بررسی فایل اکسل دسته‌جمعی، قبل از صف‌شدن در
BULK_TASKS و قبل از هرگونه اتوماسیون روی سناها.

چرا این فایل لازم است؟
  فرمول‌های شرطی داخل خود اکسل (ستون «بررسی خودکار») فقط کاربر را در لحظهٔ
  تایپ راهنمایی می‌کنند، ولی:
    - کاربر ممکن است هشدار رنگی را نبیند یا نادیده بگیرد.
    - فایل ممکن است در اکسل نسخهٔ قدیمی/گوگل‌شیت باز و فرمول‌ها خراب شوند.
    - bulk_submissions.py فعلی (parse_excel_file) اصلاً به این هشدارها کاری
      ندارد و به‌جای رد کردن داده‌ی ناقص، آن را با مقادیر ساختگی مثل
      "0000000000" جایگزین می‌کند و بی‌صدا ادامه می‌دهد - دقیقاً همان چیزی که
      باعث خطای پنهان در پردازش واقعی می‌شود.
  این ماژول با همان قوانین دقیق ربات (از validators.py) هر ردیف را دوباره،
  این‌بار در سمت سرور و مستقل از اکسل، بررسی می‌کند و گزارش شفاف برمی‌گرداند.

خروجی pre_validate_bulk_file():
    {
        "total_rows":  تعداد کل ردیف‌های غیرخالی
        "valid_items": لیست آیتم‌های آماده برای پردازش (get_bulk_task_items)
        "invalid_rows": [{"row_index": N, "errors": ["...", ...]}, ...]
        "is_all_valid": True/False
    }

استفاده:
    python bulk_validation.py path/to/file.xlsx lavayeh
    python bulk_validation.py path/to/file.xlsx ezhharnameh
"""
import sys
import logging
import openpyxl

from validators import (
    validate_tracking_code,
    validate_archive_number,
    validate_national_id,
    validate_company_id,
    validate_lavayeh_title,
    LAVAYEH_TITLES,
    DECLARANT_PERSON_TYPES,
    ADDRESSEE_PERSON_TYPES,
    to_en_digits,
)

logger = logging.getLogger(__name__)

LAVAYEH_SHEET_NAME = "ثبت دسته‌جمعی لوایح"
EZHHAR_SHEET_NAME = "ثبت دسته‌جمعی اظهارنامه"
NON_DATA_SHEETS = {"راهنما", "مرجع", "مرجع_عناوین_شعب"}


def _pick_data_sheet(wb, preferred_name: str):
    """
    مهم: در قالب جدید، شیت اول («راهنما») دیگر شیت دادهٔ اصلی نیست.
    استفاده از wb.active در اینجا اشتباه است - باید صریحاً شیت درست را پیدا کرد.
    """
    if preferred_name in wb.sheetnames:
        return wb[preferred_name]
    for name in wb.sheetnames:
        if name not in NON_DATA_SHEETS:
            return wb[name]
    return wb.active


def _cell(ws, col_letter, row):
    val = ws[f"{col_letter}{row}"].value
    if val is None:
        return ""
    return str(val).strip()


# ══════════════════════════════════════════════════════════════════════
# لایحه — ستون‌بندی مطابق «ثبت_دسته_جمعی_لوایح_هوشمند.xlsx» نسخهٔ ۳
# A ردیف | B روش | C شماره‌پرونده | D ردیف‌فرعی
# E استان‌پرونده (فقط روش=شماره‌پرونده) | F استان‌شعبه (فقط روش=بایگانی)
# G نام‌شعبه (فقط روش=بایگانی) | H شماره‌بایگانی (فقط روش=بایگانی)
# I عنوان | J-M ارائه‌دهندگان ۱-۴ | N وکیل؟ | O کدملی وکیل | P متن | Q بررسی
#
# نکته: ستون E (استان پرونده) با ستون F (استان شعبه) کاملاً متفاوت است —
# اولی از لیست PROVINCES در keyboards.py (برای _select_province، لازم در
# هر دو روش)، دومی از درخت units_compact.json (فقط برای پیداکردن نام/کد
# شعبه در روش بایگانی).
# ══════════════════════════════════════════════════════════════════════
LAVAYEH_METHOD_CASE = "شماره پرونده و ردیف فرعی"
LAVAYEH_METHOD_ARCHIVE = "شعبه و شماره بایگانی"


def parse_lavayeh_excel(filepath: str):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = _pick_data_sheet(wb, LAVAYEH_SHEET_NAME)
    rows = []
    r = 2
    max_row = ws.max_row
    while r <= max_row:
        # ستون‌های A (شمارندهٔ ردیف) و Q (بررسی خودکار) را نادیده می‌گیریم -
        # هر دو فرمولی هستند و حتی برای ردیف خالی هم مقدار دارند.
        values = [ws.cell(row=r, column=c).value for c in range(2, 17)]
        if not any(v not in (None, "") for v in values):
            r += 1
            continue
        item = {
            "row_index": r - 1,
            "method": _cell(ws, "B", r),
            "case_number": to_en_digits(_cell(ws, "C", r)),
            "sub_row": _cell(ws, "D", r) or "1",
            "province": _cell(ws, "E", r),         # استان پرونده (روش شماره‌پرونده)
            "branch_province": _cell(ws, "F", r),  # استان شعبه (روش بایگانی)
            "branch": _cell(ws, "G", r),
            "archive_number": to_en_digits(_cell(ws, "H", r)),
            "title": _cell(ws, "I", r),
            "providers": [
                to_en_digits(_cell(ws, c, r))
                for c in ("J", "K", "L", "M")
                if _cell(ws, c, r)
            ],
            "has_representative": _cell(ws, "N", r),
            "representative_id": to_en_digits(_cell(ws, "O", r)),
            "text": _cell(ws, "P", r),
        }
        rows.append(item)
        r += 1
    return rows


def validate_lavayeh_row(item: dict):
    errors = []

    # --- روش شماره‌گذاری + شماره پرونده/بایگانی ---
    if item["method"] == LAVAYEH_METHOD_CASE:
        if not item["case_number"]:
            errors.append("شماره پرونده وارد نشده")
        else:
            ok, msg = validate_tracking_code(item["case_number"])
            if not ok:
                errors.append(msg)
        if not item["province"]:
            errors.append("استان پرونده انتخاب نشده (برای روش شماره پرونده الزامی است)")
    elif item["method"] == LAVAYEH_METHOD_ARCHIVE:
        if not item["branch_province"] or not item["branch"] or not item["archive_number"]:
            errors.append("استان‌شعبه/نام‌شعبه/شماره بایگانی ناقص است")
        else:
            ok, msg = validate_archive_number(item["archive_number"])
            if not ok:
                errors.append(msg)
    else:
        errors.append("روش شماره‌گذاری انتخاب نشده یا نامعتبر است")

    # --- عنوان لایحه ---
    if not item["title"]:
        errors.append("عنوان لایحه انتخاب نشده")
    else:
        ok, msg = validate_lavayeh_title(item["title"])
        if not ok:
            errors.append(msg)

    # --- ارائه‌دهندگان ---
    if not item["providers"]:
        errors.append("حداقل یک ارائه‌دهنده الزامی است")
    for i, pid in enumerate(item["providers"], start=1):
        ok, msg = validate_national_id(pid)
        if not ok:
            errors.append(f"کدملی ارائه‌دهنده نفر {i}: {msg}")

    # --- وکیل/نماینده ---
    if item["has_representative"] == "بله" and not item["representative_id"]:
        errors.append("«آیا وکیل/نماینده دارد؟» = بله ولی کدملی وکیل وارد نشده")
    if item["title"] == "اعلام وکالت" and not item["representative_id"]:
        errors.append("برای عنوان «اعلام وکالت» کدملی وکیل الزامی است")
    if item["representative_id"]:
        ok, msg = validate_national_id(item["representative_id"])
        if not ok:
            errors.append(f"کدملی وکیل/نماینده: {msg}")

    # --- تکراری‌نبودن کدملی‌ها (ارائه‌دهندگان + وکیل) ---
    all_ids = list(item["providers"])
    if item["representative_id"]:
        all_ids.append(item["representative_id"])
    seen = set()
    for pid in all_ids:
        if pid in seen:
            errors.append(f"کدملی {pid} تکراری است")
        seen.add(pid)

    # --- متن ---
    if not item["text"]:
        errors.append("متن لایحه خالی است")

    return errors


# ══════════════════════════════════════════════════════════════════════
# اظهارنامه — ستون‌بندی مطابق «ثبت_دسته_جمعی_اظهارنامه_هوشمند.xlsx»
# اظهارکننده ۱-۴: (نوع، کدملی/شناسه، کدملی نماینده)  -> ستون‌های B..M
# مخاطب ۱-۴: (نوع، کدملی/شناسه)                      -> ستون‌های N..U  [بدون نماینده]
# V,W نمایندهٔ پرونده ۱-۲ | X عنوان | Y متن
# ══════════════════════════════════════════════════════════════════════
_DECL_SLOT_COLS = [("B", "C", "D"), ("E", "F", "G"), ("H", "I", "J"), ("K", "L", "M")]
_ADDR_SLOT_COLS = [("N", "O"), ("P", "Q"), ("R", "S"), ("T", "U")]
_REP_COLS = ["V", "W"]


def parse_ezhharnameh_excel(filepath: str):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = _pick_data_sheet(wb, EZHHAR_SHEET_NAME)
    rows = []
    r = 2
    max_row = ws.max_row
    while r <= max_row:
        # ستون‌های A (شمارندهٔ ردیف) و Z (بررسی خودکار - ستون فرمولی که حتی
        # برای ردیف خالی هم متن خطا تولید می‌کند) را نادیده می‌گیریم.
        values = [ws.cell(row=r, column=c).value for c in range(2, 26)]
        if not any(v not in (None, "") for v in values):
            r += 1
            continue

        declarants = []
        for tcol, icol, rcol in _DECL_SLOT_COLS:
            ptype = _cell(ws, tcol, r)
            pid = to_en_digits(_cell(ws, icol, r))
            rep = to_en_digits(_cell(ws, rcol, r))
            if ptype or pid:
                declarants.append({"type": ptype, "id": pid, "company_rep": rep})

        addressees = []
        for tcol, icol in _ADDR_SLOT_COLS:
            ptype = _cell(ws, tcol, r)
            pid = to_en_digits(_cell(ws, icol, r))
            if ptype or pid:
                addressees.append({"type": ptype, "id": pid})

        representatives = [
            to_en_digits(_cell(ws, c, r)) for c in _REP_COLS if _cell(ws, c, r)
        ]

        item = {
            "row_index": r - 1,
            "declarants": declarants,
            "addressees": addressees,
            "representatives": representatives,
            "title": _cell(ws, "X", r),
            "text": _cell(ws, "Y", r),
        }
        rows.append(item)
        r += 1
    return rows


def _validate_person_slot(person: dict, label: str, allowed_types, require_rep_if_legal: bool):
    """برمی‌گرداند: (errors:list, ids_to_check_for_duplicate:list)"""
    errors = []
    ids = []
    ptype = person.get("type", "")
    pid = person.get("id", "")
    rep = person.get("company_rep", "")

    if not ptype and not pid:
        return errors, ids  # اسلات استفاده‌نشده
    if not ptype:
        errors.append(f"نوع {label} انتخاب نشده")
        return errors, ids
    if ptype not in allowed_types:
        errors.append(f"نوع {label} نامعتبر است (باید یکی از: {', '.join(allowed_types)})")
        return errors, ids

    if ptype == "شخص حقوقی":
        ok, msg = validate_company_id(pid)
        if not ok:
            errors.append(f"شناسه ملی {label}: {msg}")
        else:
            ids.append(pid)
        if require_rep_if_legal:
            if not rep:
                errors.append(f"کدملی نمایندهٔ شرکت {label} الزامی است")
            else:
                ok, msg = validate_national_id(rep)
                if not ok:
                    errors.append(f"کدملی نمایندهٔ شرکت {label}: {msg}")
                else:
                    ids.append(rep)
    else:  # شخص حقیقی / وکیل
        ok, msg = validate_national_id(pid)
        if not ok:
            errors.append(f"کدملی {label}: {msg}")
        else:
            ids.append(pid)
    return errors, ids


def validate_ezhharnameh_row(item: dict):
    errors = []
    all_ids = []

    if not item["declarants"]:
        errors.append("حداقل یک اظهارکننده (نفر ۱) الزامی است")
    if not item["addressees"]:
        errors.append("حداقل یک مخاطب (نفر ۱) الزامی است")

    for i, d in enumerate(item["declarants"], start=1):
        errs, ids = _validate_person_slot(
            d, f"اظهارکننده نفر {i}", DECLARANT_PERSON_TYPES, require_rep_if_legal=True
        )
        errors += errs
        all_ids += ids

    for i, a in enumerate(item["addressees"], start=1):
        errs, ids = _validate_person_slot(
            a, f"مخاطب نفر {i}", ADDRESSEE_PERSON_TYPES, require_rep_if_legal=False
        )
        errors += errs
        all_ids += ids

    for i, rid in enumerate(item["representatives"], start=1):
        ok, msg = validate_national_id(rid)
        if not ok:
            errors.append(f"کدملی نمایندهٔ پرونده نفر {i}: {msg}")
        else:
            all_ids.append(rid)

    # قانون: اگر وکیل در اظهارکنندگان هست، باید حداقل یک حقیقی/حقوقی هم باشد
    decl_types = [d.get("type") for d in item["declarants"]]
    if "وکیل" in decl_types and not any(t in ("شخص حقیقی", "شخص حقوقی") for t in decl_types):
        errors.append("چون وکیل اضافه شده، باید حداقل یک شخص حقیقی یا حقوقی هم در اظهارکنندگان باشد")

    # تکراری‌نبودن همهٔ کدها/شناسه‌ها/نمایندگان در کل ردیف
    seen = set()
    for pid in all_ids:
        if pid in seen:
            errors.append(f"کد {pid} تکراری است")
        seen.add(pid)

    if not item["text"]:
        errors.append("متن اظهارنامه خالی است")

    return errors


# ══════════════════════════════════════════════════════════════════════
# API عمومی
# ══════════════════════════════════════════════════════════════════════
def pre_validate_bulk_file(filepath: str, service_type: str) -> dict:
    """
    service_type: "lavayeh" یا "ezhharnameh"
    """
    if service_type == "lavayeh":
        raw_rows = parse_lavayeh_excel(filepath)
        validator = validate_lavayeh_row
    elif service_type == "ezhharnameh":
        raw_rows = parse_ezhharnameh_excel(filepath)
        validator = validate_ezhharnameh_row
    else:
        raise ValueError(f"service_type نامعتبر: {service_type}")

    valid_items = []
    invalid_rows = []
    for item in raw_rows:
        errs = validator(item)
        if errs:
            invalid_rows.append({"row_index": item["row_index"], "errors": errs})
        else:
            item["status"] = "pending"
            valid_items.append(item)

    return {
        "total_rows": len(raw_rows),
        "valid_items": valid_items,
        "invalid_rows": invalid_rows,
        "is_all_valid": len(invalid_rows) == 0 and len(raw_rows) > 0,
    }


def format_report_fa(result: dict, service_type: str, max_chunk_chars: int = 3500):
    """
    گزارش فارسیِ آماده برای ارسال در بله. چون پیام‌های بله/تلگرام محدودیت طول
    دارند، خروجی را به چند بخش (chunk) تقسیم می‌کند تا هرکدام جداگانه
    با bot.send_message ارسال شود.
    برمی‌گرداند: لیستی از رشته‌ها (هرکدام یک پیام).
    """
    service_fa = "لایحه" if service_type == "lavayeh" else "اظهارنامه"
    total = result["total_rows"]
    n_invalid = len(result["invalid_rows"])
    n_valid = len(result["valid_items"])

    header = (
        f"📋 گزارش پیش‌بررسی فایل {service_fa}\n\n"
        f"تعداد کل ردیف‌ها: {total}\n"
        f"✅ سالم و آمادهٔ پردازش: {n_valid}\n"
        f"❌ دارای خطا: {n_invalid}\n"
    )

    if total == 0:
        header += "\nهیچ ردیف داده‌ای در فایل پیدا نشد. لطفاً از ردیف ۲ به بعد اطلاعات را وارد کنید."
        return [header]

    if n_invalid == 0:
        header += "\nهمهٔ ردیف‌ها معتبر هستند و می‌توانید ادامه دهید."
        return [header]

    header += "\nلطفاً خطاهای زیر را در فایل اصلاح و دوباره ارسال کنید:\n"

    chunks = [header]
    current = ""
    for row in result["invalid_rows"]:
        line = f"\n🔸 ردیف {row['row_index']}:\n"
        for e in row["errors"]:
            line += f"   • {e}\n"
        if len(current) + len(line) > max_chunk_chars:
            chunks.append(current)
            current = line
        else:
            current += line
    if current:
        chunks.append(current)
    return chunks


# ══════════════════════════════════════════════════════════════════════
# اجرای مستقل برای تست محلی (مطابق روال همیشگی هادی: تست لوکال قبل از
# اتصال به ربات)
# ══════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("استفاده: python bulk_validation.py <path.xlsx> <lavayeh|ezhharnameh>")
        sys.exit(1)

    path, stype = sys.argv[1], sys.argv[2]
    result = pre_validate_bulk_file(path, stype)
    for chunk in format_report_fa(result, stype):
        print(chunk)
        print("─" * 60)
