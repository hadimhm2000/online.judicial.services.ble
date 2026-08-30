# -*- coding: utf-8 -*-
"""
check_bulk_validation.py
──────────────────────────────────────────────────────────────────────────
معادل bulk_validation.py برای «ثبت دسته‌جمعی چک»، هم‌راستا با فایل جدید
«ثبت_دسته_جمعی_چک_هوشمند.xlsx».

چرا فایل جداگانه (و نه اضافه‌کردن مستقیم به bulk_validation.py)؟
  فقط برای اینکه هادی راحت‌تر بتواند دیف بگیرد؛ اما بهتر است محتوای این فایل
  در نهایت داخل همان bulk_validation.py (کنار توابع لایحه/اظهارنامه) ادغام
  شود تا همه‌چیز یک‌جا باشد. کافی است این فایل را کپی و importها را با
  ابتدای bulk_validation.py یکی کنید.

ستون‌بندی (دقیقاً مطابق شیت «ثبت دسته‌جمعی چک» - نسخهٔ ۲ با نوع نماینده):
  A ردیف (نادیده گرفته می‌شود)
  B عنوان خواسته | C مبلغ چک | D کدرهگیری چک
  E-T خواهان ۱-۴  → هر نفر: (نوع, کدملی/شناسه, کدملی نمایندهٔ شرکت, نوع نماینده)
  U-AJ خوانده ۱-۴ → همان ۴ ستون
  AK-AL مطلع/گواه ۱-۲ (فقط کدملی)
  AM نام شعبه (صلاحیت دادگاه) — resolve از units_output.csv (فهرست تعاملی check_branch_callback)
  AN عنوان خواسته (متن کوتاه - اختیاری) | AO شرح خواسته | AP سایر دلایل
  AQ بررسی خودکار (فرمولی - نادیده گرفته می‌شود)
  AR-AU ستون‌های کمکی فرمولی (نادیده گرفته می‌شوند)

استفاده:
    python check_bulk_validation.py path/to/file.xlsx
"""
import sys
import logging
import openpyxl

from validators import (
    validate_national_id,
    validate_company_id,
    to_en_digits,
)
from check_branches_lookup import resolve_check_branch

# resolve_check_branch از check_branches_lookup.py استفاده مجدد می‌شود — دقیقاً
# منبع درستِ صلاحیت دادگاه برای بخش چک (units_output.csv، همان لیستی که
# check_handlers.py::check_branch_callback از طریق branches.py/ROOT_NODES
# برای انتخاب تعاملی صلاحیت دادگاه نشان می‌دهد).
# ⚠️ توجه: این عمداً از bulk_submissions._resolve_branch_code (که روی
# branch_code_lookup.json کار می‌کند و مخصوص روش «بایگانی» در ثبت دسته‌جمعی
# لوایح است) استفاده نمی‌کند — آن دیتاست برای صلاحیت دادگاه چک نادرست است.

logger = logging.getLogger(__name__)

CHECK_SHEET_NAME = "ثبت دسته‌جمعی چک"
NON_DATA_SHEETS = {"راهنما", "مرجع"}

CHECK_TITLES = ("صدور اجرائیه چک", "مطالبه وجه چک")
PLAINTIFF_PERSON_TYPES = ("شخص حقیقی", "شخص حقوقی", "وکیل")
DEFENDANT_PERSON_TYPES = ("شخص حقیقی", "شخص حقوقی")

# متن‌های پیش‌فرض «عنوان خواسته» — دقیقاً همان‌هایی که check_handlers.py::check_amount_handler
# به‌عنوان متن پیشنهادی نشان می‌دهد؛ اگر ستون AF (عنوان خواسته - متن کوتاه) در اکسل خالی
# بماند، همین‌ها استفاده می‌شوند تا رفتار دسته‌جمعی با ثبت تکی یکسان بماند.
_SUGGESTED_KHASTEH_TEXT = {
    "صدور اجرائیه چک": (
        "به موجب یک فقره چک به شماره ... مورخ ... به عهده بانک ملی "
        "به مبلغ ... ریال با کدرهگیری ... به انضمام کلیه خسارات دادرسی و حق الوکاله وکیل "
        "و خسارات تاخيرتاديه از زمان سررسيد لغايت زمان كامل اجراي حكم و حق الوكاله وكيل"
    ),
    "مطالبه وجه چک": (
        "به موجب ........ فقره چک به شماره ......... مورخ ......... به عهده بانک ....... "
        "به انضمام کلیه هزینه های دادرسی و خسارات تاخیرتادیه از زمان سررسید "
        "لغایت زمان کامل اجرای حکم و حق الوکاله وکیل"
    ),
}

# ستون‌ها به‌صورت (نوع، کد/شناسه، نمایندهٔ شرکت، نوع نماینده) به ترتیب حروف اکسل
_PLAINTIFF_SLOT_COLS = [("E", "F", "G", "H"), ("I", "J", "K", "L"), ("M", "N", "O", "P"), ("Q", "R", "S", "T")]
_DEFENDANT_SLOT_COLS = [("U", "V", "W", "X"), ("Y", "Z", "AA", "AB"), ("AC", "AD", "AE", "AF"), ("AG", "AH", "AI", "AJ")]
_WITNESS_COLS = ["AK", "AL"]
_COL_BRANCH = "AM"
_COL_KHASTEH_TEXT = "AN"
_COL_TEXT = "AO"
_COL_OTHER = "AP"
REPRESENTATIVE_TYPES = ("مدیرعامل", "نماینده")


def _pick_data_sheet(wb):
    if CHECK_SHEET_NAME in wb.sheetnames:
        return wb[CHECK_SHEET_NAME]
    for name in wb.sheetnames:
        if name not in NON_DATA_SHEETS:
            return wb[name]
    return wb.active


def _cell(ws, col_letter, row):
    val = ws[f"{col_letter}{row}"].value
    if val is None:
        return ""
    return str(val).strip()


def parse_check_excel(filepath: str):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = _pick_data_sheet(wb)
    rows = []
    r = 2
    max_row = ws.max_row
    while r <= max_row:
        # ستون A (شمارندهٔ ردیف) و AQ..AU (فرمولی/بررسی خودکار) نادیده گرفته می‌شوند
        values = [ws.cell(row=r, column=c).value for c in range(2, 43)]  # B..AP
        if not any(v not in (None, "") for v in values):
            r += 1
            continue

        plaintiffs = []
        for tcol, icol, rcol, rtcol in _PLAINTIFF_SLOT_COLS:
            ptype = _cell(ws, tcol, r)
            pid = to_en_digits(_cell(ws, icol, r))
            rep = to_en_digits(_cell(ws, rcol, r))
            rep_type = _cell(ws, rtcol, r)
            if ptype or pid:
                plaintiffs.append({"type": ptype, "id": pid, "company_rep": rep, "company_rep_type": rep_type})

        defendants = []
        for tcol, icol, rcol, rtcol in _DEFENDANT_SLOT_COLS:
            ptype = _cell(ws, tcol, r)
            pid = to_en_digits(_cell(ws, icol, r))
            rep = to_en_digits(_cell(ws, rcol, r))
            rep_type = _cell(ws, rtcol, r)
            if ptype or pid:
                defendants.append({"type": ptype, "id": pid, "company_rep": rep, "company_rep_type": rep_type})

        witnesses = [
            to_en_digits(_cell(ws, c, r)) for c in _WITNESS_COLS if _cell(ws, c, r)
        ]

        item = {
            "row_index": r - 1,
            "title": _cell(ws, "B", r),
            "amount": to_en_digits(_cell(ws, "C", r)),
            "tracking_code": to_en_digits(_cell(ws, "D", r)),
            "plaintiffs": plaintiffs,
            "defendants": defendants,
            "witnesses": witnesses,
            "branch_name": _cell(ws, _COL_BRANCH, r),
            "khasteh_text": _cell(ws, _COL_KHASTEH_TEXT, r),
            "text": _cell(ws, _COL_TEXT, r),
            "extra_text": _cell(ws, _COL_OTHER, r),
        }
        rows.append(item)
        r += 1
    return rows


def _validate_person_slot(person: dict, label: str, allowed_types, require_rep_if_legal: bool):
    """برمی‌گرداند: (errors:list, ids_to_check_for_duplicate:list)"""
    errors = []
    ids = []
    ptype = person.get("type", "")
    pid = person.get("id", "")
    rep = person.get("company_rep", "")
    rep_type = person.get("company_rep_type", "")

    if not ptype and not pid:
        return errors, ids  # اسلات استفاده‌نشده
    if not ptype:
        errors.append(f"نوع {label} انتخاب نشده")
        return errors, ids
    if ptype not in allowed_types:
        errors.append(f"نوع {label} نامعتبر است (باید یکی از: {', '.join(allowed_types)})")
        return errors, ids

    if ptype == "شخص حقوقی":
        ok, msg = validate_company_id(pid)
        if not ok:
            errors.append(f"شناسه ملی {label}: {msg}")
        else:
            ids.append(pid)
        if require_rep_if_legal:
            if not rep:
                errors.append(f"کدملی نمایندهٔ شرکت {label} الزامی است")
            else:
                ok, msg = validate_national_id(rep)
                if not ok:
                    errors.append(f"کدملی نمایندهٔ شرکت {label}: {msg}")
                else:
                    ids.append(rep)
            if rep_type not in REPRESENTATIVE_TYPES:
                errors.append(
                    f"نوع نمایندهٔ شرکت {label} انتخاب نشده یا نامعتبر است "
                    f"(باید یکی از: {', '.join(REPRESENTATIVE_TYPES)})"
                )
    else:  # شخص حقیقی / وکیل
        ok, msg = validate_national_id(pid)
        if not ok:
            errors.append(f"کدملی {label}: {msg}")
        else:
            ids.append(pid)
    return errors, ids


def validate_check_row(item: dict):
    errors = []
    all_ids = []

    # --- عنوان خواسته ---
    if item["title"] not in CHECK_TITLES:
        errors.append("عنوان خواسته انتخاب نشده یا نامعتبر است")

    # --- مبلغ چک ---
    amount = item["amount"]
    if not amount or not amount.isdigit() or int(amount) <= 0:
        errors.append("مبلغ چک نامعتبر است (باید عدد مثبت به ریال باشد)")

    # --- کدرهگیری چک ---
    if not item["tracking_code"] or not item["tracking_code"].isdigit():
        errors.append("کدرهگیری چک وارد نشده یا نامعتبر است")

    # --- خواهان‌ها ---
    if not item["plaintiffs"]:
        errors.append("حداقل یک خواهان (نفر ۱) الزامی است")
    for i, p in enumerate(item["plaintiffs"], start=1):
        errs, ids = _validate_person_slot(
            p, f"خواهان نفر {i}", PLAINTIFF_PERSON_TYPES, require_rep_if_legal=True
        )
        errors += errs
        all_ids += ids

    # --- خوانده‌ها ---
    if not item["defendants"]:
        errors.append("حداقل یک خوانده (نفر ۱) الزامی است")
    for i, d in enumerate(item["defendants"], start=1):
        errs, ids = _validate_person_slot(
            d, f"خوانده نفر {i}", DEFENDANT_PERSON_TYPES, require_rep_if_legal=True
        )
        errors += errs
        all_ids += ids

    # --- مطلع/گواه ---
    for i, wid in enumerate(item["witnesses"], start=1):
        ok, msg = validate_national_id(wid)
        if not ok:
            errors.append(f"کدملی مطلع/گواه نفر {i}: {msg}")
        else:
            all_ids.append(wid)

    # --- تکراری‌نبودن همهٔ کدها/شناسه‌ها/نمایندگان در کل ردیف ---
    seen = set()
    for pid in all_ids:
        if pid in seen:
            errors.append(f"کد {pid} تکراری است")
        seen.add(pid)

    # --- نام شعبه (صلاحیت دادگاه) ---
    if not item["branch_name"]:
        errors.append("نام شعبه (صلاحیت دادگاه) وارد نشده")

    # --- شرح خواسته ---
    if not item["text"]:
        errors.append("شرح خواسته (متن) خالی است")

    return errors


# ══════════════════════════════════════════════════════════════════════
# API عمومی — امضا دقیقاً مثل pre_validate_bulk_file در bulk_validation.py
# ══════════════════════════════════════════════════════════════════════
def pre_validate_check_bulk_file(filepath: str) -> dict:
    raw_rows = parse_check_excel(filepath)

    valid_items = []
    invalid_rows = []
    for item in raw_rows:
        errs = validate_check_row(item)

        branch_code = ""
        if not errs:
            branch_code, branch_item, suggestions = resolve_check_branch(item["branch_name"])
            if not branch_code:
                msg = f"نام شعبه «{item['branch_name']}» در فهرست واحدهای قضائی پیدا نشد."
                if suggestions:
                    msg += " آیا منظورتان یکی از این‌ها بود؟ " + " / ".join(suggestions)
                errs = errs + [msg]
            else:
                # نام/مسیر را با نسخهٔ دقیق و رسمی (از units_output.csv) جایگزین می‌کنیم
                # تا آنچه به check_scenario.py می‌رود همیشه دقیقاً با چیزی که در سنا
                # ثبت می‌شود یکی باشد (نه املای دستیِ کاربر در اکسل).
                item["branch_name"] = branch_item["name"]
                item["branch_path"] = branch_item["path"]

        if errs:
            invalid_rows.append({"row_index": item["row_index"], "errors": errs})
        else:
            item["status"] = "pending"
            item["branch_code"] = branch_code
            if not item["khasteh_text"]:
                item["khasteh_text"] = _SUGGESTED_KHASTEH_TEXT.get(item["title"], "")
            item["check_images"] = []          # تصاویر چک همین ردیف - تا تکمیل مرحلهٔ پیوست خالی می‌ماند
            item["has_check_image_title"] = False  # آیا حداقل یک‌بار «🧾 تصویر چک» انتخاب شده
            valid_items.append(item)

    return {
        "total_rows": len(raw_rows),
        "valid_items": valid_items,
        "invalid_rows": invalid_rows,
        "is_all_valid": len(invalid_rows) == 0 and len(raw_rows) > 0,
    }


def format_check_report_fa(result: dict, max_chunk_chars: int = 3500):
    total = result["total_rows"]
    n_invalid = len(result["invalid_rows"])
    n_valid = len(result["valid_items"])

    header = (
        f"📋 گزارش پیش‌بررسی فایل دعاوی چک\n\n"
        f"تعداد کل ردیف‌ها: {total}\n"
        f"✅ سالم و آمادهٔ پردازش: {n_valid}\n"
        f"❌ دارای خطا: {n_invalid}\n"
    )

    if total == 0:
        header += "\nهیچ ردیف داده‌ای در فایل پیدا نشد. لطفاً از ردیف ۲ به بعد اطلاعات را وارد کنید."
        return [header]

    if n_invalid == 0:
        header += (
            "\nهمهٔ ردیف‌ها معتبر هستند. حالا برای هر ردیف باید حداقل یک‌بار "
            "«🧾 تصویر چک» را انتخاب و ۳ تصویر ضمیمه کنید تا ارسال نهایی انجام شود."
        )
        return [header]

    header += "\nلطفاً خطاهای زیر را در فایل اصلاح و دوباره ارسال کنید:\n"

    chunks = [header]
    current = ""
    for row in result["invalid_rows"]:
        line = f"\n🔸 ردیف {row['row_index']}:\n"
        for e in row["errors"]:
            line += f"   • {e}\n"
        if len(current) + len(line) > max_chunk_chars:
            chunks.append(current)
            current = line
        else:
            current += line
    if current:
        chunks.append(current)
    return chunks


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("استفاده: python check_bulk_validation.py <path.xlsx>")
        sys.exit(1)

    result = pre_validate_check_bulk_file(sys.argv[1])
    for chunk in format_check_report_fa(result):
        print(chunk)
        print("─" * 60)
