# -*- coding: utf-8 -*-
"""
bulk_inquiry_excel.py
──────────────────────────────────────────────────────────────────────────
استعلام دسته‌جمعی از طریق فایل اکسل (کدرهگیری / شماره موبایل / کدملی).

قالب: templates/sample_bulk_inquiry.xlsx — ستون‌بندی (۵ ستون):
    A کدرهگیری
    B دریافت پیوست‌ها؟ (بله/خیر)
    C نوع سند — شامل زیرشاخه‌ها به‌صورت فلت (مثلاً «دعاوی اعتراضی - تجدیدنظرخواهی»)
    D شماره موبایل
    E کدملی

نکات نسبت به نسخه‌ی قبلی:
  • ستون جداگانه‌ی «زیرشاخه» حذف شد؛ برای انواع سندی که زیرشاخه دارند،
    هر زیرشاخه یک آیتم مستقل در همان دراپ‌داون «نوع سند» است
    (با جداکننده‌ی " - ")؛ CATEGORY_OPTIONS پایین لیست کامل را نگه می‌دارد
    و split_category() آن را به (category, subcategory) تجزیه می‌کند.
  • اعتبارسنجی/ساخت آیتم per-column است.
  • اگر در میانه‌ی پردازش «نشست سامانه منقضی» تشخیص داده شود، بلافاصله
    حلقه متوقف می‌شود (به‌جای اینکه تک‌تک تمام ردیف‌های باقی‌مانده را هم
    امتحان کند و همان خطا را ده‌ها بار تکرار کند).
  • خروجی cart_items دقیقاً هم‌ساختار آیتم سبد خرید موجود است
    (query_type / tracking_code / doc_category / doc_subcategory /
    need_attachments / fee / total_attachments).
"""
import logging
import re
import openpyxl

from config import FEES, get_fee

logger = logging.getLogger(__name__)

TRACKING_CODE_PREFIX_MIN = 1394220
TRACKING_CODE_PREFIX_MAX = 1406220
TRACKING_CODE_LENGTH = 16
ATTACHMENT_PAGE_RATE = 5000  # تومان به ازای هر برگ پیوست — هم‌راستا با handlers.py

DATA_SHEET_NAME = "استعلام دسته‌جمعی"
NON_DATA_SHEETS = {"راهنما", "لیست‌ها"}

# ── لیست کامل «نوع سند» — دسته‌های ساده + زیرشاخه‌های فلت‌شده ─────────
SIMPLE_CATEGORIES = [
    "لایحه", "اظهارنامه", "شکواییه", "دادخواست بدوی", "دعاوی دادگاههای صلح",
]
SUBCATEGORY_MAP = {
    "دعاوی اعتراضی": [
        "تجدیدنظرخواهی", "واخواهی", "فرجام خواهی", "اعاده دادرسی مدنی",
        "اعاده دادرسی کیفری", "اعتراض ثالث", "اعتراض به قرار دادسرا",
    ],
    "دعاوی طاری": ["دعوای تقابل", "دعوای ورود ثالث", "دعوای جلب ثالث"],
    "شورای حل اختلاف": [
        "دعاوی حقوقی", "دعاوی کیفری", "تجدیدنظرخواهی شورا",
        "واخواهی شورا", "اعتراض ثالث شورا",
    ],
    "دیوان عدالت اداری": [
        "دادخواست بدوی دیوان عدالت اداری", "تجدیدنظرخواهی دیوان عدالت اداری",
        "ارایه و پیگیری لایحه", "جلب ثالث در بدوی دیوان عدالت اداری",
        "ورود ثالث در بدوی دیوان عدالت اداری",
    ],
}
CATEGORY_SEP = " - "


def build_flat_category_options():
    """تولید لیست کامل گزینه‌های دراپ‌داون «نوع سند» (فلت‌شده)."""
    options = list(SIMPLE_CATEGORIES)
    for cat, subs in SUBCATEGORY_MAP.items():
        for sub in subs:
            options.append(f"{cat}{CATEGORY_SEP}{sub}")
    return options


CATEGORY_OPTIONS = build_flat_category_options()


def split_category(selected: str):
    """تجزیه‌ی مقدار انتخاب‌شده از دراپ‌داون به (category, subcategory).
    اگر زیرشاخه نداشت subcategory=None برمی‌گردد."""
    if not selected:
        return None, None
    if CATEGORY_SEP in selected:
        cat, sub = selected.split(CATEGORY_SEP, 1)
        return cat.strip(), sub.strip()
    return selected.strip(), None


# ══════════════════════════════════════════════════════════════════════
def to_en_digits(s) -> str:
    """تبدیل ارقام فارسی/عربی به انگلیسی. اگر مقدار float صحیح بود، بدون
    اعشار به رشته تبدیل می‌شود تا صفر ابتدایی/دقت رقمی خراب نشود."""
    if s is None:
        return ""
    if isinstance(s, float) and s.is_integer():
        s = str(int(s))
    else:
        s = str(s)
    table = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")
    return s.translate(table).replace(" ", "").replace("‌", "").strip()


def extract_digits(raw) -> str:
    """استخراج فقط ارقام از مقدار سلول — هر حرف/خط‌تیره/فاصله (مثل پیشوند
    اجباری T/M/N که در راهنما خواسته‌ایم) نادیده گرفته می‌شود.

    این تابع دلیل اصلی رفع باگ «کدرهگیری/موبایل به عدد تبدیل و رقم آخر یا
    صفر ابتدایی گم می‌شود» است: چون کاربر دیگر عدد خالص در سلول تایپ
    نمی‌کند (بلکه چیزی مثل T1405220948201280 یا M09123456789)، هیچ
    اپلیکیشن اکسل/گوگل‌شیتی (حتی پیش‌نمایش‌های محدود) این را به‌عنوان عدد
    تشخیص نمی‌دهد و همیشه دقیقاً همان‌طور که تایپ شده ذخیره می‌ماند."""
    s = to_en_digits(raw)
    return re.sub(r"\D", "", s)


def _fix_leading_zero(value: str, expected_len: int) -> str:
    """اگر با وجود همه‌ی این تدابیر باز هم صفر ابتدایی افتاده باشد (مثلاً
    فایل قدیمی بدون پیشوند حرفی)، دوباره با صفر به طول موردنظر می‌رساند.
    فقط برای موبایل/کدملی — نه کدرهگیری."""
    if value and value.isdigit() and len(value) < expected_len:
        return value.zfill(expected_len)
    return value


def _is_valid_tracking_code(code: str) -> bool:
    if len(code) != TRACKING_CODE_LENGTH or not code.isdigit():
        return False
    prefix = int(code[:7])
    return TRACKING_CODE_PREFIX_MIN <= prefix <= TRACKING_CODE_PREFIX_MAX


def _pick_data_sheet(wb):
    if DATA_SHEET_NAME in wb.sheetnames:
        return wb[DATA_SHEET_NAME]
    for name in wb.sheetnames:
        if name not in NON_DATA_SHEETS:
            return wb[name]
    return wb.active


def _cell(ws, col_letter, row):
    val = ws[f"{col_letter}{row}"].value
    if val is None:
        return ""
    return str(val).strip()


# ══════════════════════════════════════════════════════════════════════
# مرحله ۱: خواندن خام (۵ ستون: A تا E)
# ══════════════════════════════════════════════════════════════════════
def parse_bulk_inquiry_excel(filepath: str):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = _pick_data_sheet(wb)
    rows = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, 6)]
        if not any(v not in (None, "") for v in values):
            continue
        row = {
            "row_index": r - 1,
            "tracking_code": extract_digits(_cell(ws, "A", r)),
            "need_attachments_raw": _cell(ws, "B", r),
            "doc_category_selected": _cell(ws, "C", r),
            "phone": _fix_leading_zero(extract_digits(_cell(ws, "D", r)), 11),
            "national_id": _fix_leading_zero(extract_digits(_cell(ws, "E", r)), 10),
        }
        rows.append(row)
    return rows


# ══════════════════════════════════════════════════════════════════════
# مرحله ۲: اعتبارسنجی/ساخت آیتم — PER-COLUMN
# ══════════════════════════════════════════════════════════════════════
def build_bulk_inquiry_items(filepath: str):
    raw_rows = parse_bulk_inquiry_excel(filepath)
    valid_items = []
    invalid_cells = []

    for row in raw_rows:
        row_idx = row["row_index"]
        tracking_code = row["tracking_code"]
        phone = row["phone"]
        national_id = row["national_id"]

        # --- کدرهگیری ---
        if tracking_code:
            if not _is_valid_tracking_code(tracking_code):
                invalid_cells.append({
                    "row_index": row_idx, "field": "کدرهگیری",
                    "error": f"کدرهگیری «{tracking_code}» نامعتبر است "
                             f"(باید دقیقاً {TRACKING_CODE_LENGTH} رقم و در بازه‌ی مجاز باشد؛ "
                             f"اگر ستون به‌صورت عدد نمایش داده می‌شود، حتماً قبل از خودِ کد یک حرف "
                             f"بگذارید، مثلاً T1405220948201280)",
                })
            else:
                need_att_raw = (row.get("need_attachments_raw") or "").strip()
                need_attachments = need_att_raw == "بله"
                selected = row.get("doc_category_selected") or ""
                doc_category, doc_subcategory = split_category(selected)

                if need_attachments and not doc_category:
                    invalid_cells.append({
                        "row_index": row_idx, "field": "نوع سند",
                        "error": f"کدرهگیری «{tracking_code}»: برای محاسبه‌ی تعداد پیوست، "
                                 f"ستون «نوع سند» باید از لیست کشویی انتخاب شود",
                    })
                else:
                    valid_items.append({
                        "row_index": row_idx,
                        "kind": "tracking",
                        "tracking_code": tracking_code,
                        "need_attachments": need_attachments,
                        "doc_category": doc_category,
                        "doc_subcategory": doc_subcategory,
                    })

        # --- شماره موبایل ---
        if phone:
            if not (len(phone) == 11 and phone.isdigit() and phone.startswith("09")):
                invalid_cells.append({
                    "row_index": row_idx, "field": "شماره موبایل",
                    "error": f"شماره موبایل «{phone}» نامعتبر است (باید ۱۱ رقم و با 09 شروع شود)",
                })
            else:
                valid_items.append({"row_index": row_idx, "kind": "phone", "phone": phone})

        # --- کدملی ---
        if national_id:
            if not (len(national_id) == 10 and national_id.isdigit()):
                invalid_cells.append({
                    "row_index": row_idx, "field": "کدملی",
                    "error": f"کدملی «{national_id}» نامعتبر است (باید دقیقاً ۱۰ رقم باشد)",
                })
            else:
                valid_items.append({"row_index": row_idx, "kind": "national_id", "national_id": national_id})

    return {
        "total_rows": len(raw_rows),
        "valid_items": valid_items,
        "invalid_cells": invalid_cells,
    }


# ══════════════════════════════════════════════════════════════════════
# مرحله ۳: تکمیل هزینه — یک fast_pre_check دقیق برای هر کدرهگیریِ
# نیازمند پیوست. اگر نشست سامانه منقضی باشد، بلافاصله متوقف می‌شود.
# ══════════════════════════════════════════════════════════════════════
async def enrich_bulk_inquiry_items(valid_items: list, user_id: int = None, bot=None):
    from api_direct import (
        fast_pre_check, FastCheckError,
        SessionExpiredError as FastSessionExpiredError,
        PetitionNotFoundError as FastPetitionNotFoundError,
        InvalidTrackingCodeError as FastInvalidTrackingCodeError,
    )
    import runtime_state

    cart_items = []
    failed_items = []
    session_broken = False

    for it in valid_items:
        kind = it["kind"]
        row_idx = it["row_index"]

        if kind == "phone":
            cart_items.append({
                "query_type": "شماره تماس", "tracking_code": it["phone"],
                "doc_category": None, "doc_subcategory": None,
                "need_attachments": False, "fee": get_fee("شماره تماس", False),
                "total_attachments": 0, "row_index": row_idx,
            })
            continue

        if kind == "national_id":
            cart_items.append({
                "query_type": "کد ملی", "tracking_code": it["national_id"],
                "doc_category": None, "doc_subcategory": None,
                "need_attachments": False, "fee": get_fee("کد ملی", False),
                "total_attachments": 0, "row_index": row_idx,
            })
            continue

        # kind == "tracking"
        tracking_code = it["tracking_code"]
        need_attachments = it["need_attachments"]
        doc_category = it["doc_category"]
        doc_subcategory = it["doc_subcategory"]

        if not need_attachments:
            cart_items.append({
                "query_type": "کد رهگیری", "tracking_code": tracking_code,
                "doc_category": doc_category, "doc_subcategory": doc_subcategory,
                "need_attachments": False, "fee": get_fee("کد رهگیری", False),
                "total_attachments": 0, "row_index": row_idx,
            })
            continue

        if session_broken:
            # نشست سامانه در یکی از موارد قبلی منقضی شد — دیگر لازم نیست
            # همین خطا را برای بقیه‌ی ردیف‌ها هم تکرار کنیم.
            failed_items.append({
                "row_index": row_idx,
                "description": f"کدرهگیری {tracking_code} ({doc_category})",
                "error": "پردازش نشد (نشست سامانه سناخا در حین کار قطع شد)",
            })
            continue

        if not getattr(runtime_state, "browser_context", None):
            session_broken = True
            failed_items.append({
                "row_index": row_idx,
                "description": f"کدرهگیری {tracking_code} ({doc_category})",
                "error": "مرورگر/نشست سامانه سناخا آماده نیست — به مدیر اطلاع دهید",
            })
            continue

        try:
            total_attachments_count = await fast_pre_check(
                tracking_code=tracking_code, category=doc_category,
                subcategory=doc_subcategory, user_id=user_id, bot=bot,
            )
            fee = FEES["کد رهگیری با منضمات"] + total_attachments_count * ATTACHMENT_PAGE_RATE
            cart_items.append({
                "query_type": "کد رهگیری", "tracking_code": tracking_code,
                "doc_category": doc_category, "doc_subcategory": doc_subcategory,
                "need_attachments": True, "fee": fee,
                "total_attachments": total_attachments_count, "row_index": row_idx,
            })
        except FastPetitionNotFoundError:
            failed_items.append({
                "row_index": row_idx,
                "description": f"کدرهگیری {tracking_code} ({doc_category})",
                "error": "پرونده‌ای با این کد رهگیری/نوع سند یافت نشد",
            })
        except FastInvalidTrackingCodeError:
            failed_items.append({
                "row_index": row_idx,
                "description": f"کدرهگیری {tracking_code} ({doc_category})",
                "error": "کدرهگیری یا نوع سند اشتباه است",
            })
        except FastSessionExpiredError as e:
            session_broken = True
            failed_items.append({
                "row_index": row_idx,
                "description": f"کدرهگیری {tracking_code} ({doc_category})",
                "error": f"نشست سامانه سناخا منقضی است ({e}) — بقیه‌ی موارد هم متوقف شدند",
            })
        except FastCheckError as e:
            failed_items.append({
                "row_index": row_idx,
                "description": f"کدرهگیری {tracking_code} ({doc_category})",
                "error": f"خطا در استعلام از سامانه ({e})",
            })
        except Exception as e:
            logger.error(f"[BULK-INQ] خطای غیرمنتظره برای {tracking_code}: {e}", exc_info=True)
            failed_items.append({
                "row_index": row_idx,
                "description": f"کدرهگیری {tracking_code} ({doc_category})",
                "error": "خطای غیرمنتظره در پردازش",
            })

    return cart_items, failed_items


# ══════════════════════════════════════════════════════════════════════
# مرحله ۴: فاکتور
# ══════════════════════════════════════════════════════════════════════
def build_invoice_text(cart_items: list, invalid_cells: list = None, failed_items: list = None):
    invalid_cells = invalid_cells or []
    failed_items = failed_items or []

    n_tracking = sum(1 for i in cart_items if i["query_type"] == "کد رهگیری")
    n_phone = sum(1 for i in cart_items if i["query_type"] == "شماره تماس")
    n_national = sum(1 for i in cart_items if i["query_type"] == "کد ملی")

    lines = ["🧾 *فاکتور استعلام دسته‌جمعی*", ""]
    lines.append(
        f"📦 تعداد کل موارد معتبر: *{len(cart_items)} مورد* "
        f"(کدرهگیری: {n_tracking} | موبایل: {n_phone} | کدملی: {n_national})"
    )
    lines.append("")
    lines.append("📋 *جزئیات ردیف‌به‌ردیف:*")

    total_sum = 0
    for idx, item in enumerate(sorted(cart_items, key=lambda x: x["row_index"]), start=1):
        row_no = item["row_index"]
        fee = item["fee"]
        total_sum += fee

        if item["query_type"] == "شماره تماس":
            desc = f"ردیف {row_no} — 📞 استعلام موبایل `{item['tracking_code']}`"
            calc = f"{fee:,} تومان (نرخ ثابت)"
        elif item["query_type"] == "کد ملی":
            desc = f"ردیف {row_no} — 👤 استعلام کدملی `{item['tracking_code']}`"
            calc = f"{fee:,} تومان (نرخ ثابت)"
        else:
            cat_label = item['doc_category']
            if item.get('doc_subcategory'):
                cat_label += f" - {item['doc_subcategory']}"
            if item["need_attachments"]:
                pages = item["total_attachments"]
                base = FEES["کد رهگیری با منضمات"]
                desc = (
                    f"ردیف {row_no} — 📄 کدرهگیری `{item['tracking_code']}` "
                    f"({cat_label}) + پیوست ({pages} برگ)"
                )
                calc = f"{base:,} + ({pages} × {ATTACHMENT_PAGE_RATE:,}) = *{fee:,} تومان*"
            else:
                desc = f"ردیف {row_no} — 📄 کدرهگیری `{item['tracking_code']}` (بدون پیوست)"
                calc = f"{fee:,} تومان (نرخ ثابت)"

        lines.append(f"{idx}. {desc}\n   💰 {calc}")

    lines.append("")
    lines.append(f"💰 *مجموع کل: {total_sum:,} تومان*")

    if invalid_cells:
        lines.append("")
        lines.append(f"⚠️ *{len(invalid_cells)} خطای فرمت — از فاکتور کنار گذاشته شد:*")
        for c in invalid_cells[:15]:
            lines.append(f"  • ردیف {c['row_index']} ({c['field']}): {c['error']}")
        if len(invalid_cells) > 15:
            lines.append(f"  ... و {len(invalid_cells) - 15} خطای دیگر")

    if failed_items:
        lines.append("")
        lines.append(f"⚠️ *{len(failed_items)} مورد در استعلام از سامانه ناموفق بود:*")
        for f in failed_items[:15]:
            lines.append(f"  • ردیف {f['row_index']} ({f['description']}): {f['error']}")

    return "\n".join(lines), total_sum
