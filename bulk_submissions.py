# -*- coding: utf-8 -*-
"""
ماژول پردازش ثبت‌های دسته‌جمعی (بیش از ۵ مورد)
شامل:
- پارسر منعطف اکسل (قالب هوشمند جدید + قدیمی)
- صف پردازش واقعی با ارسال به job_queue
"""

import os
import re
import json
import random
import string
import asyncio
import logging
from datetime import datetime
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
from states import Form
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

BULK_TASKS = {}

# ══════════════════════════════════════════════════════════════════════
# نگاشت «نام کامل شعبه» -> «کد ۵ رقمی شعبه» (#txtCourtCode)
# ══════════════════════════════════════════════════════════════════════
_BRANCH_CODE_LOOKUP_PATH = os.path.join(os.path.dirname(__file__), "branch_code_lookup.json")
_branch_code_cache = None


def _load_branch_code_lookup() -> dict:
    """بارگذاری و کش کردن branch_code_lookup.json"""
    global _branch_code_cache
    if _branch_code_cache is None:
        try:
            with open(_BRANCH_CODE_LOOKUP_PATH, encoding="utf-8") as f:
                _branch_code_cache = json.load(f)
            logger.info(f"[BULK] branch_code_lookup.json بارگذاری شد ({len(_branch_code_cache)} ورودی)")
        except FileNotFoundError:
            logger.error(f"[BULK] branch_code_lookup.json پیدا نشد در {_BRANCH_CODE_LOOKUP_PATH}")
            _branch_code_cache = {}
        except Exception as e:
            logger.error(f"[BULK] خطا در بارگذاری branch_code_lookup.json: {e}")
            _branch_code_cache = {}
    return _branch_code_cache


def _resolve_branch_code(branch_name: str) -> str:
    """استخراج کد ۵ رقمی شعبه از نام شعبه با استفاده از branch_code_lookup.json.

    جستجو در چند قالب انجام می‌شود:
    1. تطبیق دقیق نام کامل
    2. جستجوی زیررشته (اگر نام شعبه بخشی از کلید باشد)
    3. جستجوی معکوس (اگر کلید بخشی از نام شعبه باشد)
    """
    if not branch_name:
        return ""

    lookup = _load_branch_code_lookup()
    branch_name = branch_name.strip()

    # ۱) تطبیق دقیق
    if branch_name in lookup:
        return lookup[branch_name]

    # ۲) جستجوی زیررشته - نام شعبه بخشی از کلید باشد
    for key, code in lookup.items():
        if branch_name in key or key in branch_name:
            logger.info(f"[BULK] تطبیق شعبه: '{branch_name}' -> '{key}' (کد: {code})")
            return code

    logger.warning(f"[BULK] کد شعبه برای '{branch_name}' پیدا نشد")
    return ""


def generate_tracking_code(prefix="BLK") -> str:
    digits = ''.join(random.choices(string.digits, k=6))
    return f"#{prefix}-{digits}"


def generate_sample_excel(service_type: str, filepath: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    if service_type == "ezhharnameh":
        ws.title = "نمونه ثبت دسته‌جمعی اظهارنامه"
    else:
        ws.title = "نمونه ثبت دسته‌جمعی لوایح"
    ws.views.sheetView[0].rightToLeft = True

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style='thin', color='CCCCCC')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    if service_type == "lavayeh":
        headers = ["ردیف", "شماره پرونده 16 یا 18 رقمی", "ردیف فرعی",
                   "عنوان لایحه (از لیست عناوین را انتخاب کنید)",
                   "کدملی شخص ارائه دهنده",
                   "متن لایحه (متن را کاملا کپی کنید)"]
    else:
        headers = ["ردیف", "کدملی / شناسه ملی اظهارکننده",
                   "کدملی / شناسه ملی مخاطب",
                   "کدملی نماینده (درصورت وجود)",
                   "عنوان اظهارنامه (خالی=سایر)",
                   "متن اظهارنامه (متن را کاملا کپی کنید)"]

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = cell_border

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath


def _pick_data_sheet(wb, service_type: str):
    if service_type == "lavayeh":
        preferred = ["ثبت دسته‌جمعی لوایح"]
    else:
        preferred = ["ثبت دسته‌جمعی اظهارنامه"]
    skip = {"راهنما", "مرجع", "مرجع_عناوین_شعب"}
    for name in preferred:
        for sn in wb.sheetnames:
            if sn.replace("\u200c", "") == name.replace("\u200c", ""):
                return wb[sn]
    for sn in wb.sheetnames:
        if sn not in skip:
            return wb[sn]
    return wb.active


def _to_en_digits(text: str) -> str:
    fa = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")
    return text.translate(fa).replace(" ", "").strip()


def _is_empty_row_lavayeh(ws, r: int, max_col: int = 16) -> bool:
    check_up_to = max(7, min(max_col, ws.max_column))
    for c in range(2, check_up_to + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != "":
            return False
    return True


def _detect_lavayeh_format(ws) -> str:
    header_b = str(ws.cell(row=1, column=2).value or "").strip()
    if "\u0631\u0648\u0634" in header_b:
        return "smart"
    return "legacy"


def _is_empty_row_ezhharnameh(ws, r: int, max_col: int = 25) -> bool:
    check_up_to = max(7, min(max_col, ws.max_column))
    for c in range(2, check_up_to + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != "":
            return False
    return True


def _detect_ezhharnameh_format(ws) -> str:
    header_b = str(ws.cell(row=1, column=2).value or "").strip()
    if "\u0646\u0648\u0639 \u0627\u0638\u0647\u0627\u0631\u06a9\u0646\u0646\u062f\u0647" in header_b:
        return "smart"
    return "legacy"


def _cell_value(ws, col: int, row: int) -> str:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    raw = str(v).strip()
    # حذف کاراکترهای نامعتبر یونیکد از اکسل
    return ''.join(c for c in raw if not (0xD800 <= ord(c) <= 0xDFFF) and ord(c) != 0xFFFD)


def parse_excel_file(filepath: str, service_type: str) -> dict:
    valid_items = []
    invalid_rows = []
    total_rows = 0

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = _pick_data_sheet(wb, service_type)
        max_row = ws.max_row

        is_empty = _is_empty_row_lavayeh if service_type == "lavayeh" else _is_empty_row_ezhharnameh

        LAVAYEH_TITLES_SET = {
            "\u0644\u0627\u06cc\u062d\u0647 \u062f\u0641\u0627\u0639\u06cc\u0647", "\u0635\u062f\u0648\u0631 \u0627\u062c\u0631\u0627\u0626\u06cc\u0647", "\u0627\u0639\u062a\u0631\u0627\u0636 \u0628\u0647 \u0646\u0638\u0631 \u06a9\u0627\u0631\u0634\u0646\u0627\u0633",
            "\u0627\u0639\u062a\u0631\u0627\u0636 \u0628\u0647 \u0642\u0631\u0627\u0631 \u0631\u062f \u062f\u0641\u062a\u0631", "\u0627\u0639\u0644\u0627\u0645 \u0648\u06a9\u0627\u0644\u062a",
            "\u062f\u0631\u062e\u0648\u0627\u0633\u062a \u0645\u0645\u0646\u0648\u0639\u06cc\u062a \u0627\u0632 \u062e\u0631\u0648\u062c \u06a9\u0634\u0648\u0631", "\u062f\u0631\u062e\u0648\u0627\u0633\u062a \u06a9\u067e\u06cc \u0627\u0632 \u0645\u062f\u0627\u0631\u06a9 \u067e\u0631\u0648\u0646\u062f\u0647",
            "\u062f\u0631\u062e\u0648\u0627\u0633\u062a \u0645\u0637\u0627\u0644\u0628\u0647 \u067e\u0631\u0648\u0646\u062f\u0647", "\u062f\u0631\u062e\u0648\u0627\u0633\u062a \u0645\u0637\u0627\u0644\u0639\u0647 \u067e\u0631\u0648\u0646\u062f\u0647", "\u0633\u0627\u06cc\u0631 \u0639\u0646\u0627\u0648\u06cc\u0646"
        }

        consecutive_empty = 0
        MAX_CONSECUTIVE_EMPTY = 5

        r = 2
        while r <= max_row:
            if is_empty(ws, r):
                consecutive_empty += 1
                if consecutive_empty >= MAX_CONSECUTIVE_EMPTY:
                    break
                r += 1
                continue

            consecutive_empty = 0
            total_rows += 1
            row_num = r - 1
            errors = []

            if service_type == "lavayeh":
                fmt = _detect_lavayeh_format(ws)

                if fmt == "smart":
                    method = _cell_value(ws, 2, r)
                    case_number = _to_en_digits(_cell_value(ws, 3, r))
                    sub_row = _to_en_digits(_cell_value(ws, 4, r)) or "1"
                    province_case = _cell_value(ws, 5, r)
                    province_branch = _cell_value(ws, 6, r)
                    branch_name = _cell_value(ws, 7, r)
                    archive_number = _to_en_digits(_cell_value(ws, 8, r))
                    title = _cell_value(ws, 9, r)
                    text = _cell_value(ws, 16, r)

                    providers = []
                    for pc in [10, 11, 12, 13]:
                        pid = _to_en_digits(_cell_value(ws, pc, r))
                        if pid:
                            providers.append(pid)

                    has_lawyer = _cell_value(ws, 14, r)
                    lawyer_id = _to_en_digits(_cell_value(ws, 15, r))

                    method_clean = ""
                    if "\u0628\u0627\u06cc\u06af\u0627\u0646\u06cc" in method:
                        method_clean = "\u0628\u0627\u06cc\u06af\u0627\u0646\u06cc"
                    elif "\u0634\u0645\u0627\u0631\u0647 \u067e\u0631\u0648\u0646\u062f\u0647" in method or "\u067e\u0631\u0648\u0646\u062f\u0647" in method:
                        method_clean = "\u0634\u0645\u0627\u0631\u0647 \u067e\u0631\u0648\u0646\u062f\u0647"

                    if not method_clean:
                        errors.append("\u0631\u0648\u0634 \u0634\u0645\u0627\u0631\u0647\u200c\u06af\u0630\u0627\u0631\u06cc \u0627\u0646\u062a\u062e\u0627\u0628 \u0646\u0634\u062f\u0647")
                    if method_clean == "\u0634\u0645\u0627\u0631\u0647 \u067e\u0631\u0648\u0646\u062f\u0647" and not case_number:
                        errors.append("\u0634\u0645\u0627\u0631\u0647 \u067e\u0631\u0648\u0646\u062f\u0647 \u0648\u0627\u0631\u062f \u0646\u0634\u062f\u0647")
                    if method_clean == "\u0628\u0627\u06cc\u06af\u0627\u0646\u06cc" and not archive_number:
                        errors.append("\u0634\u0645\u0627\u0631\u0647 \u0628\u0627\u06cc\u06af\u0627\u0646\u06cc \u0648\u0627\u0631\u062f \u0646\u0634\u062f\u0647")
                    if not title or title == "\u0639\u0646\u0648\u0627\u0646 \u0644\u0627\u06cc\u062d\u0647 \u0627\u0646\u062a\u062e\u0627\u0628 \u0646\u0634\u062f\u0647":
                        errors.append("\u0639\u0646\u0648\u0627\u0646 \u0644\u0627\u06cc\u062d\u0647 \u0627\u0646\u062a\u062e\u0627\u0628 \u0646\u0634\u062f\u0647")
                    elif title not in LAVAYEH_TITLES_SET:
                        errors.append(f"\u0639\u0646\u0648\u0627\u0646 \u0644\u0627\u06cc\u062d\u0647 \u00ab{title}\u00bb \u0646\u0627\u0645\u0639\u062a\u0628\u0631 \u0627\u0633\u062a")
                    if not providers:
                        errors.append("\u06a9\u062f\u0645\u0644\u06cc \u0627\u0631\u0627\u0626\u0647\u200c\u062f\u0647\u0646\u062f\u0647 \u0627\u0644\u0632\u0627\u0645\u06cc \u0627\u0633\u062a")
                    if not text:
                        errors.append("\u0645\u062a\u0646 \u0644\u0627\u06cc\u062d\u0647 \u062e\u0627\u0644\u06cc \u0627\u0633\u062a")

                    if errors:
                        invalid_rows.append({"row_index": row_num, "errors": errors})
                    else:
                        item = {
                            "row_index": row_num, "method": method_clean, "title": title,
                            "providers": providers, "text": text,
                            "attachments": [], "status": "pending"
                        }
                        if method_clean == "\u0634\u0645\u0627\u0631\u0647 \u067e\u0631\u0648\u0646\u062f\u0647":
                            item["case_number"] = case_number
                            item["sub_row"] = sub_row
                            item["province"] = province_case
                        elif method_clean == "\u0628\u0627\u06cc\u06af\u0627\u0646\u06cc":
                            item["archive_number"] = archive_number
                            item["province"] = province_branch
                            item["branch_name"] = branch_name
                        if has_lawyer and "\u0628\u0644\u0647" in has_lawyer and lawyer_id:
                            item["lawyer_id"] = lawyer_id
                        valid_items.append(item)

                else:
                    case_number = _to_en_digits(_cell_value(ws, 2, r))
                    sub_row = _to_en_digits(_cell_value(ws, 3, r)) or "1"
                    title = _cell_value(ws, 4, r)
                    provider_id = _to_en_digits(_cell_value(ws, 5, r))
                    text = _cell_value(ws, 6, r)
                    if not case_number: errors.append("\u0634\u0645\u0627\u0631\u0647 \u067e\u0631\u0648\u0646\u062f\u0647 \u0648\u0627\u0631\u062f \u0646\u0634\u062f\u0647")
                    if not title: errors.append("\u0639\u0646\u0648\u0627\u0646 \u0644\u0627\u06cc\u062d\u0647 \u0627\u0646\u062a\u062e\u0627\u0628 \u0646\u0634\u062f\u0647")
                    elif title not in LAVAYEH_TITLES_SET:
                        errors.append(f"\u0639\u0646\u0648\u0627\u0646 \u0644\u0627\u06cc\u062d\u0647 \u00ab{title}\u00bb \u0646\u0627\u0645\u0639\u062a\u0628\u0631 \u0627\u0633\u062a")
                    if not provider_id: errors.append("\u06a9\u062f\u0645\u0644\u06cc \u0627\u0631\u0627\u0626\u0647\u200c\u062f\u0647\u0646\u062f\u0647 \u0627\u0644\u0632\u0627\u0645\u06cc \u0627\u0633\u062a")
                    if not text: errors.append("\u0645\u062a\u0646 \u0644\u0627\u06cc\u062d\u0647 \u062e\u0627\u0644\u06cc \u0627\u0633\u062a")
                    if errors:
                        invalid_rows.append({"row_index": row_num, "errors": errors})
                    else:
                        valid_items.append({"row_index": row_num, "case_number": case_number, "sub_row": sub_row,
                            "title": title, "providers": [provider_id], "text": text,
                            "attachments": [], "status": "pending"})

            else:
                fmt = _detect_ezhharnameh_format(ws)

                if fmt == "smart":
                    declarants = []
                    for gs in [2, 5, 8, 11]:
                        dtype = _cell_value(ws, gs, r)
                        did = _to_en_digits(_cell_value(ws, gs + 1, r))
                        drep = _to_en_digits(_cell_value(ws, gs + 2, r))
                        if did:
                            if "\u062d\u0642\u0648\u0642\u06cc" in dtype: dtype_clean = "\u062d\u0642\u0648\u0642\u06cc"
                            elif "\u0648\u06a9\u06cc\u0644" in dtype: dtype_clean = "\u0648\u06a9\u06cc\u0644"
                            elif "\u062d\u0642\u06cc\u0642\u06cc" in dtype: dtype_clean = "\u062d\u0642\u06cc\u0642\u06cc"
                            else: dtype_clean = "\u062d\u0642\u06cc\u0642\u06cc"
                            declarants.append({"type": dtype_clean, "id": did, "company_rep": drep})

                    addressees = []
                    for gs in [14, 16, 18, 20]:
                        atype = _cell_value(ws, gs, r)
                        aid = _to_en_digits(_cell_value(ws, gs + 1, r))
                        if aid:
                            if "\u062d\u0642\u0648\u0642\u06cc" in atype:
                                atype_clean = "\u062d\u0642\u0648\u0642\u06cc"
                            elif "\u062d\u0642\u06cc\u0642\u06cc" in atype:
                                atype_clean = "\u062d\u0642\u06cc\u0642\u06cc"
                            else:
                                atype_clean = "\u062d\u0642\u06cc\u0642\u06cc"
                            addressees.append({"type": atype_clean, "id": aid})

                    representatives = []
                    for rc in [22, 23]:
                        rep_id = _to_en_digits(_cell_value(ws, rc, r))
                        if rep_id: representatives.append(rep_id)

                    title = _cell_value(ws, 24, r) or "\u0633\u0627\u06cc\u0631"
                    text = _cell_value(ws, 25, r)

                    if not declarants: errors.append("\u0627\u0638\u0647\u0627\u0631\u06a9\u0646\u0646\u062f\u0647 \u0648\u0627\u0631\u062f \u0646\u0634\u062f\u0647")
                    has_lawyer = any(d["type"] == "\u0648\u06a9\u06cc\u0644" for d in declarants)
                    has_non_lawyer = any(d["type"] != "\u0648\u06a9\u06cc\u0644" for d in declarants)
                    if has_lawyer and not has_non_lawyer:
                        errors.append("\u0627\u0638\u0647\u0627\u0631\u06a9\u0646\u0646\u062f\u0647 \u0648\u06a9\u06cc\u0644 \u062f\u0627\u0631\u062f \u0627\u0645\u0627 \u062d\u0642\u06cc\u0642\u06cc/\u062d\u0642\u0648\u0642\u06cc \u0646\u06cc\u0632 \u0644\u0627\u0632\u0645 \u0627\u0633\u062a")
                    if not addressees: errors.append("\u0645\u062e\u0627\u0637\u0628 \u0648\u0627\u0631\u062f \u0646\u0634\u062f\u0647")
                    if not text: errors.append("\u0645\u062a\u0646 \u0627\u0638\u0647\u0627\u0631\u0646\u0627\u0645\u0647 \u062e\u0627\u0644\u06cc \u0627\u0633\u062a")

                    if errors:
                        invalid_rows.append({"row_index": row_num, "errors": errors})
                    else:
                        valid_items.append({"row_index": row_num, "declarants": declarants,
                            "addressees": addressees, "representatives": representatives,
                            "title": title, "text": text, "attachments": [], "status": "pending"})

                else:
                    declarant_id = _to_en_digits(_cell_value(ws, 2, r))
                    addressee_id = _to_en_digits(_cell_value(ws, 3, r))
                    representative_id = _to_en_digits(_cell_value(ws, 4, r))
                    title = _cell_value(ws, 5, r) or "\u0633\u0627\u06cc\u0631"
                    text = _cell_value(ws, 6, r)
                    if not declarant_id: errors.append("\u06a9\u062f\u0645\u0644\u06cc \u0627\u0638\u0647\u0627\u0631\u06a9\u0646\u0646\u062f\u0647 \u0648\u0627\u0631\u062f \u0646\u0634\u062f\u0647")
                    if not addressee_id: errors.append("\u06a9\u062f\u0645\u0644\u06cc \u0645\u062e\u0627\u0637\u0628 \u0648\u0627\u0631\u062f \u0646\u0634\u062f\u0647")
                    if not text: errors.append("\u0645\u062a\u0646 \u0627\u0638\u0647\u0627\u0631\u0646\u0627\u0645\u0647 \u062e\u0627\u0644\u06cc \u0627\u0633\u062a")
                    if errors:
                        invalid_rows.append({"row_index": row_num, "errors": errors})
                    else:
                        valid_items.append({"row_index": row_num,
                            "declarants": [{"type": "", "id": declarant_id, "company_rep": representative_id}],
                            "addressees": [{"type": "", "id": addressee_id}],
                            "representatives": [representative_id] if representative_id else [],
                            "title": title, "text": text, "attachments": [], "status": "pending"})

            r += 1

    except Exception as e:
        logger.error(f"Error parsing Excel file {filepath}: {e}", exc_info=True)

    return {"valid_items": valid_items, "invalid_rows": invalid_rows, "total_rows": total_rows}


def parse_text_or_image_input(raw_text: str, service_type: str) -> list:
    lines = [l.strip() for l in raw_text.split("\n") if l.strip()]
    items = []
    for idx, line in enumerate(lines, start=1):
        if service_type == "lavayeh":
            items.append({"row_index": idx, "tracking_code": f"AUTO-{idx:03d}",
                "row_number": "1", "branch_name": "\u062b\u0628\u062a \u062f\u0633\u062a\u0647\u200c\u062c\u0645\u0639\u06cc",
                "title": "\u0644\u0627\u06cc\u062d\u0647 \u062f\u0641\u0627\u0639\u06cc\u0647 (\u0648\u0631\u0648\u062f \u0633\u0631\u06cc\u0639)",
                "national_id": "0000000000", "text": line, "attachment": "\u0646\u062f\u0627\u0631\u062f", "status": "pending"})
        else:
            items.append({"row_index": idx, "declarant_id": "0000000000",
                "addressee_id": "0000000000", "subject": "\u0627\u0638\u0647\u0627\u0631\u0646\u0627\u0645\u0647 (\u0648\u0631\u0648\u062f \u0633\u0631\u06cc\u0639)",
                "text": line, "attachment": "\u0646\u062f\u0627\u0631\u062f", "status": "pending"})
    return items


def _sanitize_text(text: str) -> str:
    """حذف کاراکترهای نامعتبر یونیکد (surrogate) از متن."""
    if not text:
        return ""
    try:
        # ابتدا encode با surrogatepass تا سوروگیت‌ها قابل شناسایی شوند
        encoded = text.encode('utf-8', errors='surrogatepass')
        # سپس decode عادی — هر کاراکتر نامعتبری حذف می‌شود
        cleaned = encoded.decode('utf-8', errors='replace')
        # حذف کاراکتر جایگزین و سوروگیت‌های باقی‌مانده
        cleaned = ''.join(c for c in cleaned if c != '\ufffd' and not (0xD800 <= ord(c) <= 0xDFFF))
        return cleaned
    except Exception:
        # فال‌بک: حذف ساده
        return ''.join(c for c in text if not (0xD800 <= ord(c) <= 0xDFFF) and c != '\ufffd')


def _transform_ezhhar_declarants(declarants: list) -> list:
    """
    تبدیل فرمت پارسر اکسل به فرمت مورد انتظار browser_worker.
    دقیقاً مشابه ساختار FSM فلوی ثبت تکی:
      حقیقی: {"person_type": "شخص حقیقی", "national_id": "..."}
      حقوقی: {"person_type": "شخص حقوقی", "company_id": "...", "representative_type": "مدیرعامل", "national_id": "..."}
      وکیل:   {"person_type": "وکیل", "national_id": "..."}
    """
    result = []
    for d in declarants:
        dtype = d.get("type", "حقیقی")
        if "وکیل" in dtype:
            person_type = "وکیل"
        elif "حقوقی" in dtype:
            person_type = "شخص حقوقی"
        else:
            person_type = "شخص حقیقی"

        if person_type == "شخص حقوقی":
            company_rep = d.get("company_rep", "")
            if company_rep:
                transformed = {
                    "person_type": "شخص حقوقی",
                    "company_id": d.get("id", ""),
                    "representative_type": "مدیرعامل",
                    "national_id": company_rep,
                }
            else:
                # بدون نماینده — شبیه addressee حقوقی
                transformed = {
                    "person_type": "شخص حقوقی",
                    "company_id": d.get("id", ""),
                    "representative_type": "",
                    "national_id": "",
                }
        else:
            transformed = {"person_type": person_type, "national_id": d.get("id", "")}

        logger.info(f"[BULK-TRANSFORM] اظهارکننده: type={repr(dtype)} -> {transformed}")
        result.append(transformed)
    return result


def _transform_ezhhar_addressees(addressees: list) -> list:
    """
    تبدیل فرمت پارسر اکسل به فرمت مورد انتظار browser_worker.
    دقیقاً مشابه ساختار FSM فلوی ثبت تکی:
      حقیقی: {"person_type": "شخص حقیقی", "national_id": "..."}
      حقوقی: {"person_type": "شخص حقوقی", "company_id": "...", "representative_type": "", "national_id": ""}
    """
    result = []
    for a in addressees:
        atype = a.get("type", "حقیقی")
        if "حقوقی" in atype:
            person_type = "شخص حقوقی"
        else:
            person_type = "شخص حقیقی"

        if person_type == "شخص حقوقی":
            # دقیقاً مشابه ezhharnameh_handlers خط 363-365:
            transformed = {
                "person_type": "شخص حقوقی",
                "company_id": a.get("id", ""),
                "representative_type": "",
                "national_id": "",
            }
        else:
            transformed = {"person_type": "شخص حقیقی", "national_id": a.get("id", "")}

        logger.info(f"[BULK-TRANSFORM] مخاطب: {transformed}")
        result.append(transformed)
    return result


def _safe_send(bot, user_id: int, text: str, parse_mode: str = None):
    """ارسال امن پیام بدون خطای یونیکد."""
    try:
        kwargs = {"chat_id": user_id, "text": _sanitize_text(text)}
        if parse_mode:
            kwargs["parse_mode"] = parse_mode
        return bot.bot.session.coro_make_request('sendMessage', kwargs)
    except Exception as e:
        logger.error(f"خطا در ارسال پیام: {e}")
        try:
            kwargs = {"chat_id": user_id, "text": _sanitize_text(text)}
            return bot.bot.session.coro_make_request('sendMessage', kwargs)
        except Exception as e2:
            logger.error(f"خطا مجدد در ارسال پیام: {e2}")


def count_processable_items(items: list, service_type: str) -> int:
    """شمارش ردیف‌های واقعاً قابل‌پردازش (پس از حذف نامعتبرها).
    
    ردیف‌های دارای status="pending" شمرده می‌شوند.
    این تابع باید هم در bulk_confirm_handler (برای محاسبه مبلغ پیش‌پرداخت)
    و هم در run_bulk_processing_task (برای فیلتر واقعی) صدا زده شود.
    """
    count = 0
    for item in items:
        if item.get("status") == "pending":
            # فیلتر عناوین پشتیبانی‌نشده برای لایحه
            if service_type == "lavayeh":
                title = item.get("title", "")
                if title in ("اعلام وکالت",):
                    continue
            count += 1
    return count


def format_bulk_tracking_list(batch_tracking_code: str) -> str:
    """تولید متن وضعیت لحظه‌ای تمام آیتم‌های یک دسته جمعی."""
    task_data = BULK_TASKS.get(batch_tracking_code)
    if not task_data:
        return "\u26a0\ufe0f کد رهگیری یافت نشد."
    
    items = task_data.get("signable_items", [])
    if not items:
        return f"\U0001f4c1 لیست خالی است (کد: `{batch_tracking_code}`)"
    
    from datetime import datetime, timedelta
    now = datetime.now()
    lines = [f"\U0001f4cb *لیست کدهای رهگیری (`{batch_tracking_code}`)*\n"]
    
    awaiting = []
    signed = []
    disabled = []
    failed_list = []
    
    for item in items:
        status = item.get("status", "unknown")
        tc = item.get("tracking_code", "?")
        title = item.get("title", "?")
        court = item.get("court_total", 0)
        
        if status == "signed":
            signed.append(f"  \u2705 `{tc}` — {title} ({court:,} ریال)")
        elif status == "failed":
            failed_list.append(f"  \u274c `{tc}` — {title} — {item.get('error_summary', 'خطا')}")
        elif status == "disabled":
            disabled_until = item.get("disabled_until")
            if disabled_until:
                if isinstance(disabled_until, str):
                    try:
                        disabled_until = datetime.fromisoformat(disabled_until)
                    except Exception:
                        disabled_until = None
                if disabled_until and now < disabled_until:
                    mins = int((disabled_until - now).total_seconds() // 60)
                    disabled.append(f"  \u23f3 `{tc}` — {title} (غیرفعال تا {mins} دقیقه دیگر)")
                else:
                    # زمان منقضی شده → بازگشت به awaiting
                    item["status"] = "awaiting_sign"
                    item["disabled_until"] = None
                    awaiting.append(f"  \U0001f514 `{tc}` — {title} ({court:,} ریال)")
            else:
                awaiting.append(f"  \U0001f514 `{tc}` — {title} ({court:,} ریال)")
        else:
            awaiting.append(f"  \U0001f514 `{tc}` — {title} ({court:,} ریال)")
    
    if signed:
        lines.append(f"\u2705 *امضا شده ({len(signed)}):*\n" + "\n".join(signed))
    if awaiting:
        lines.append(f"\U0001f514 *در انتظار امضا ({len(awaiting)}):*\n" + "\n".join(awaiting))
    if disabled:
        lines.append(f"\u23f3 *موقتاً غیرفعال ({len(disabled)}):*\n" + "\n".join(disabled))
    if failed_list:
        lines.append(f"\u274c *ناموفق ({len(failed_list)}):*\n" + "\n".join(failed_list))
    
    return "\n".join(lines)


async def _show_bulk_sign_menu(bot, user_id: int, batch_tracking_code: str):
    """نمایش منوی انتخاب کد رهگیری برای امضای دسته‌جمعی."""
    task_data = BULK_TASKS.get(batch_tracking_code)
    if not task_data:
        await _safe_send(bot, user_id, "\u26a0\ufe0f اطلاعات دسته جمعی یافت نشد.")
        return
    
    items = task_data.get("signable_items", [])
    if not items:
        await _safe_send(bot, user_id, "\U0001f4c1 لیست کدی برای امضا وجود ندارد.")
        return
    
    # بررسی قفل امضای هم‌زمان
    if task_data.get("signing_in_progress"):
        await _safe_send(bot, user_id,
            f"\u23f3 در حال امضای کد `{task_data['signing_in_progress']}` هستید.\n"
            "لطفاً ابتدا امضای جاری را تکمیل کنید.")
        return
    
    from datetime import datetime
    now = datetime.now()
    
    # ساخت inline keyboard
    buttons = []
    for item in items:
        status = item.get("status", "")
        tc = item.get("tracking_code", "")
        title = item.get("title", "?")
        
        if status == "awaiting_sign":
            # بررسی disabled_until منقضی شده
            disabled_until = item.get("disabled_until")
            if disabled_until:
                if isinstance(disabled_until, str):
                    try:
                        disabled_until = datetime.fromisoformat(disabled_until)
                    except Exception:
                        disabled_until = None
                if disabled_until and now < disabled_until:
                    mins = int((disabled_until - now).total_seconds() // 60)
                    buttons.append([InlineKeyboardButton(
                text=f"\u23f3 {tc} — {title} (غیرفعال تا {mins} دقیقه)",
                callback_data=f"bulk_sign_disabled"
            )])
            continue
        
        if status == "signed":
            buttons.append([InlineKeyboardButton(
                text=f"\u2705 {tc} — {title} (امضا شده)",
                callback_data=f"bulk_sign_disabled"
            )])
            continue
        
        if status == "failed":
            buttons.append([InlineKeyboardButton(
                text=f"\u274c {tc} — {title} (ناموفق)",
                callback_data=f"bulk_sign_disabled"
            )])
            continue
        
        # awaiting_sign بدون disabled_until منقضی
        buttons.append([InlineKeyboardButton(
            text=f"\U0001f514 {tc} — {title}",
            callback_data=f"bulk_sign_select:{tc}"
        )])
    
    # دکمه نمایش لیست کامل
    buttons.append([InlineKeyboardButton(
        text=f"\U0001f4cb نمایش کامل لیست",
        callback_data=f"bulk_sign_list:{batch_tracking_code}"
    )])
    
    from aiogram.types import InlineKeyboardMarkup
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    awaiting_count = sum(1 for i in items if i.get("status") == "awaiting_sign")
    signed_count = sum(1 for i in items if i.get("status") == "signed")
    
    msg = (
        f"\U0001f4cb *منوی امضای دسته‌جمعی*\n\n"
        f"\U0001f514 در انتظار امضا: *{awaiting_count}*\n"
        f"\u2705 امضا شده: *{signed_count}*\n\n"
        f"\U0001f4a1 یک کد رهگیری را برای ارسال کد امضا انتخاب کنید:\n"
        f"_(هر امضا = یک کد پیامکی جداگانه روی موبایل ثبت‌شده)_"
    )
    
    try:
        await bot.send_message(user_id, msg, parse_mode="Markdown", reply_markup=kb)
    except Exception as e:
        logger.error(f"خطا در ارسال منوی امضا: {e}")


async def run_bulk_processing_task(bot, user_id: int, tracking_code: str):
    """
    پردازش واقعی ثبت دسته‌جمعی — استفاده از همان توابع کمکی اصلی
    (_send_lavayeh_task_to_queue / _send_ezhhar_task_to_queue)
    تا فرمت داده‌ها دقیقاً مشابه فلوی ثبت تکی باشد.
    """
    import runtime_state

    task_data = BULK_TASKS.get(tracking_code)
    if not task_data:
        return

    items = task_data.get("items", [])
    total = len(items)
    service_type = task_data.get("service_type", "lavayeh")
    service_fa = "لایحه" if service_type == "lavayeh" else "اظهارنامه"

    task_data["status"] = "processing"
    task_data.setdefault("signable_items", [])

    # ایمپورت تنبل برای جلوگیری از ایمپورت حلقوی
    _send_lavayeh_fn = None
    _send_ezhhar_fn = None
    try:
        from lavayeh_handlers import _send_lavayeh_task_to_queue as _send_lavayeh_fn
        from ezhharnameh_handlers import _send_ezhhar_task_to_queue as _send_ezhhar_fn
    except ImportError as e:
        logger.error(f"[BULK] خطا در ایمپورت توابع ثبت اصلی: {e}")

    try:
        await bot.send_message(
            user_id,
            f"⏳ *پردازش در پس‌زمینه آغاز شد!*\n\n"
            f"کد پیگیری دسته‌جمعی: `{tracking_code}`\n"
            f"تعداد موارد: *{total} مورد ({service_fa})*\n\n"
            f"💡 موارد یکی‌یکی در سامانه ثبت خواهند شد.",
            parse_mode="Markdown")
    except Exception as e:
        logger.error(f"خطا در ارسال پیام شروع پردازش: {e}")

    # ── مجموعه جلوگیری از ثبت تکراری در سطح صف ──
    _queued_bulk_keys = set()

    queued = 0
    skipped_duplicates = 0
    errors = 0
    for idx, item in enumerate(items, start=1):
        try:
            if service_type == "ezhharnameh":
                # ساخت داده‌ها دقیقاً مشابه FSM state فلوی ثبت تکی اظهارنامه
                transformed_declarants = _transform_ezhhar_declarants(item.get("declarants", []))
                transformed_addressees = _transform_ezhhar_addressees(item.get("addressees", []))

                fsm_data = {
                    "ezhhar_declarants": transformed_declarants,
                    "ezhhar_addressees": transformed_addressees,
                    "ezhhar_subject": _sanitize_text(item.get("title", "سایر")),
                    "ezhhar_text": _sanitize_text(item.get("text", "")),
                    "ezhhar_text_html": "",
                    "ezhhar_attachments": item.get("attachments", []),
                    "_is_bulk": True,
                    "batch_tracking_code": tracking_code,
                    "_bulk_row_index": item.get("row_index", idx),
                }

                logger.info(f"[BULK-QUEUE] اظهارنامه مورد {idx}: {len(transformed_declarants)} اظهارکننده, {len(transformed_addressees)} مخاطب")
                for dd in transformed_declarants:
                    logger.info(f"  اظهارکننده: {dd}")
                for aa in transformed_addressees:
                    logger.info(f"  مخاطب: {aa}")
                logger.info(f"  عنوان: {_sanitize_text(item.get('title', 'سایر'))}")
                logger.info(f"  طول متن: {len(item.get('text', ''))}")
                logger.info(f"  تعداد پیوست: {len(item.get('attachments', []))}")
                logger.info(f"  نمایندگان پرونده: {item.get('representatives', [])}")
                logger.info(f"[BULK-QUEUE] COMPLETE JOB DICT: {fsm_data}")

                if _send_ezhhar_fn:
                    await _send_ezhhar_fn(fsm_data, user_id)
                else:
                    # فال‌بک: مستقیم در صف قرار دادن با فرمت صحیح
                    await runtime_state.job_queue.put({
                        "user_id": user_id,
                        "query_type": "اظهارنامه_ثبت",
                        "task_type": "EZHHARNAMEH_SUBMIT",
                        "ezhhar_declarants": transformed_declarants,
                        "ezhhar_addressees": transformed_addressees,
                        "ezhhar_subject": _sanitize_text(item.get("title", "سایر")),
                        "ezhhar_text": _sanitize_text(item.get("text", "")),
                        "ezhhar_text_html": "",
                        "ezhhar_attachments": item.get("attachments", []),
                        "_is_bulk": True,
                        "batch_tracking_code": tracking_code,
                        "_bulk_row_index": item.get("row_index", idx),
                    })
            else:
                method = item.get("method", "شماره پرونده")
                title = item.get("title", "لایحه دفاعیه")
                system_title = "لایحه دفاعیه" if title == "سایر عناوین" else title

                providers = item.get("providers", [])
                persons = [{"person_type": "شخص حقیقی", "national_id": _sanitize_text(pid)} for pid in providers]
                if item.get("lawyer_id"):
                    persons.append({"person_type": "وکیل", "national_id": _sanitize_text(item["lawyer_id"])})

                # شماره رهگیری واقعی = شماره پرونده یا بایگانی (نه کد بچ!)
                actual_tracking = _sanitize_text(item.get("case_number", "")) \
                    if method == "شماره پرونده" \
                    else _sanitize_text(item.get("archive_number", ""))

                # استخراج کد شعبه از نام شعبه (برای روش بایگانی)
                _branch_name_raw = item.get("branch_name", "")
                _resolved_branch_code = _resolve_branch_code(_branch_name_raw) \
                    if method == "شعبه و شماره بایگانی" else ""
                if method == "شعبه و شماره بایگانی" and _resolved_branch_code:
                    logger.info(f"[BULK-QUEUE] کد شعبه استخراج شد: '{_branch_name_raw}' -> '{_resolved_branch_code}'")
                elif method == "شعبه و شماره بایگانی":
                    logger.warning(f"[BULK-QUEUE] ⚠️ کد شعبه برای '{_branch_name_raw}' پیدا نشد!")

                # ساخت داده‌ها دقیقاً مشابه FSM state فلوی ثبت تکی لایحه
                # ══════════════════════════════════════════════════════════
                # جلوگیری از ثبت تکراری در سطح صف: اگر این مورد قبلاً
                # در صف قرار گرفته، دوباره اضافه نکن
                # ══════════════════════════════════════════════════════════
                _dup_key = f"{actual_tracking}:{item.get('sub_row', 1)}"
                if _dup_key in _queued_bulk_keys:
                    logger.warning(f"[BULK-QUEUE] ⚠️ ردیف تکراری رد شد: {_dup_key} (مورد {idx})")
                    item["status"] = "duplicate_skipped"
                    continue
                _queued_bulk_keys.add(_dup_key)

                fsm_data = {
                    "lavayeh_title": title,
                    "lavayeh_system_title": system_title,
                    "lavayeh_tracking_code": actual_tracking,
                    "lavayeh_province": _sanitize_text(item.get("province", "")),
                    "lavayeh_row_number": int(item.get("sub_row", 1) or 1),
                    "lavayeh_persons": persons,
                    "lavayeh_text": _sanitize_text(item.get("text", "")),
                    "lavayeh_text_html": "",
                    "lavayeh_attachments": item.get("attachments", []),
                    "tracking_method": "case_number" if method == "شماره پرونده" else "archive_number",
                    "lavayeh_archive_number": _sanitize_text(item.get("archive_number", "")),
                    "lavayeh_branch_name": _sanitize_text(item.get("branch_name", "")),
                    "lavayeh_branch_code": _resolved_branch_code,
                    "_is_bulk": True,
                    "batch_tracking_code": tracking_code,
                }

                if _send_lavayeh_fn:
                    await _send_lavayeh_fn(fsm_data, user_id, title)
                else:
                    # فال‌بک: مستقیم در صف قرار دادن با فرمت صحیح
                    await runtime_state.job_queue.put({
                        "user_id": user_id,
                        "query_type": "لایحه_ثبت",
                        "task_type": "LAVAYEH_SUBMIT",
                        "lavayeh_title": title,
                        "lavayeh_system_title": system_title,
                        "lavayeh_tracking_code": actual_tracking,
                        "lavayeh_province": _sanitize_text(item.get("province", "")),
                        "lavayeh_row_number": int(item.get("sub_row", 1) or 1),
                        "lavayeh_persons": persons,
                        "lavayeh_text": _sanitize_text(item.get("text", "")),
                        "lavayeh_text_html": "",
                        "lavayeh_attachments": item.get("attachments", []),
                        "tracking_method": "case_number" if method == "شماره پرونده" else "archive_number",
                        "lavayeh_archive_number": _sanitize_text(item.get("archive_number", "")),
                        "lavayeh_branch_name": _sanitize_text(item.get("branch_name", "")),
                        "lavayeh_branch_code": _resolved_branch_code,
                        "_is_bulk": True,
                    })

            if item.get("status") == "duplicate_skipped":
                skipped_duplicates += 1
                logger.warning(f"[BULK-QUEUE] مورد {idx} تکراری رد شد")
                continue

            item["status"] = "queued"
            queued += 1
            logger.info(f"[BULK-QUEUE] مورد {idx}/{total} در صف قرار گرفت (کد: {tracking_code})")

            if queued % 5 == 0 or queued == total:
                try:
                    await bot.send_message(user_id,
                        f"🔄 *گزارش پیشرفت (`{tracking_code}`)*\n\n"
                        f"📥 در صف ارسال: *{queued} از {total}*\n"
                        f"📌 ردیف {idx} به صف پردازش سامانه اضافه شد.",
                        parse_mode="Markdown")
                except Exception as e:
                    logger.error(f"خطا در ارسال گزارش پیشرفت: {e}")

        except Exception as e:
            logger.error(f"[BULK-QUEUE] خطا در ارسال مورد {idx} به صف: {e}", exc_info=True)
            item["status"] = "error"
            errors += 1

    task_data["status"] = "queued"
    task_data["queued_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    task_data["queued_count"] = queued

    dup_note = f"\n⚠️ {skipped_duplicates} مورد تکراری رد شد." if skipped_duplicates else ""
    error_note = f"\n⚠️ {errors} مورد خطا داشت." if errors else ""
    try:
        await bot.send_message(user_id,
            f"✅ *تمام {queued} مورد در صف پردازش سامانه قرار گرفت!*{error_note}{dup_note}\n\n"
            f"🔒 کد پیگیری: `{tracking_code}`\n"
            f"📥 موارد در صف: *{queued} از {total}*\n\n"
            f"⏳ موارد یکی‌یکی ثبت خواهند شد و نتیجه هر مورد برایتان ارسال می‌شود.",
            parse_mode="Markdown")
    except Exception as e:
        logger.error(f"خطا در ارسال پیام پایان: {e}")

    # ══════════════════════════════════════════════════════════
    # گزارش مالی و صدور فاکتور تسویه (در صورت نیاز)
    # ══════════════════════════════════════════════════════════
    signable_items = task_data.get("signable_items", [])
    if signable_items:
        total_court_cost = sum(item.get("court_total", 0) for item in signable_items if item.get("status") != "failed")
        prepaid_total = task_data.get("prepaid_total_rial", 0)
        remaining = total_court_cost - prepaid_total

        if remaining > 0:
            diff_line = f"\U0001f4b0 مابه‌التفاوه (باقیمانده): *{remaining:,} ریال*"
        elif remaining < 0:
            diff_line = f"\u2705 پیش‌پرداخت پوشش داده (مازاد: {abs(remaining):,} ریال)"
        else:
            diff_line = "\u2705 پیش‌پرداخت دقیقاً برابر هزینه سامانه"

        report_msg = (
            f"\U0001f9fe *گزارش مالی دسته‌جمعی (`{tracking_code}`)*\n\n"
            f"\U0001f4b0 مجموع هزینه واقعی سامانه: *{total_court_cost:,} ریال*\n"
            f"\U0001f4b3 مجموع پیش‌پرداخت شما: *{prepaid_total:,} ریال*\n"
            f"{diff_line}"
        )

        # فهرست ردیف‌های ناموفق
        failed_items = [item for item in signable_items if item.get("status") == "failed"]
        if failed_items:
            report_msg += f"\n\n\u274c *{len(failed_items)} ردیف ناموفق:*\n"
            for fi in failed_items:
                report_msg += f"  \u2022 ردیف {fi.get('row_index', '?')}: {fi.get('title', '?')} — {fi.get('error_summary', 'خطای نامشخص')}\n"

        try:
            await bot.send_message(user_id, report_msg, parse_mode="Markdown")
        except Exception as e:
            logger.error(f"خطا در ارسال گزارش مالی: {e}")

        # صدور فاکتور تسویه فقط اگر باقیمانده > 0
        if remaining > 0:
            try:
                from config import BALE_WALLET_TOKEN, BALE_API_BASE, BOT_TOKEN
                import aiohttp
                import json as _json

                invoice_payload = _json.dumps({"type": "bulk_settlement", "uid": user_id, "tracking_code": tracking_code})
                async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=False)) as session:
                    invoice_url = f"{BALE_API_BASE}/bot{BOT_TOKEN}/sendInvoice"
                    invoice_data = {
                        "chat_id": user_id,
                        "title": f"تسویه هزینه سامانه ({tracking_code})",
                        "description": f"مابه‌التفاوه هزینه سامانه قضایی\nمبلغ: {remaining // 10:,} تومان ({remaining:,} ریال)",
                        "payload": invoice_payload,
                        "provider_token": BALE_WALLET_TOKEN,
                        "currency": "IRR",
                        "prices": [{"label": f"تسویه {tracking_code}", "amount": remaining}],
                    }
                    async with session.post(invoice_url, json=invoice_data) as resp:
                        result = await resp.json()
                        if not result.get("ok"):
                            logger.error(f"[BULK-SETTLE] خطای sendInvoice: {result}")
                            raise Exception(result.get("description", "خطا در ارسال فاکتور"))

                from aiogram.types import ReplyKeyboardRemove
                await bot.send_message(user_id,
                    "\u23f3 فاکتور مابه‌التفاوه هزینه سامانه ارسال شد.\n"
                    "پس از پرداخت، منوی انتخاب کد رهگیری برای امضا نمایش داده می‌شود.",
                    reply_markup=ReplyKeyboardRemove())

                # ذخیره state برای تشخیص پس از پرداخت
                task_data["settlement_amount_rial"] = remaining
                # در اینجا باید state کاربر تنظیم شود — از طریق runtime_state
                import runtime_state as _rs
                if hasattr(_rs, 'dp') and _rs.dp:
                    user_state = _rs.dp.fsm.resolve_context(bot, user_id, user_id)
                    await user_state.set_state(Form.bulk_settlement_wait)

            except Exception as e:
                logger.error(f"[BULK-SETTLE] خطا در صدور فاکتور تسویه: {e}", exc_info=True)
                await _safe_send(bot, user_id, f"\u26a0\ufe0f خطا در صدور فاکتور تسویه. لطفاً به مدیریت اطلاع دهید.")
        else:
            # remaining <= 0 → مستقیم به منوی امضا
            await _show_bulk_sign_menu(bot, user_id, tracking_code)
    elif signable_items is not None and len(signable_items) == 0:
        # هیچ ردیفی موفق نبود
        await _safe_send(bot, user_id,
            f"\u274c هیچ ردیفی با موفقیت ثبت نشد. لطفاً خطاهای گزارش‌شده را برطرف و مجدداً تلاش فرمایید.")
